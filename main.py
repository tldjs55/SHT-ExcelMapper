#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SHT Excel CSV寫入工具
自動識別Excel中的特定欄位和空格，建立寫入
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import win32com.client


FIELD_MAPPING_DIR = os.path.expanduser("~/documents/field_mappings")
FIELD_MAPPING_PATH = os.path.join(FIELD_MAPPING_DIR, "field_mappings.json")

# UI常量
MAX_HORIZONTAL_SCAN_RANGE = 4
MAX_VERTICAL_SCAN_RANGE = 8

class SmartExcelMapper:
	"""Excel寫入工具"""

	def __init__(self):
		self.root = tk.Tk()
		self.root.title("RATC Excel Mapper (SHT)")
		self.root.geometry("1400x1000")

		# 數據存儲
		self.csv_data = []
		self.excel_data = []
		self.excel_workbook = None
		self.excel_sheet = None
		self.active_workbook = None
		self.active_worksheet = None

		# Excel連接模式
		self.auto_detect_mode = True  # 預設使用自動偵測

		# 寫入配置
		self.field_mappings = {}  # 存儲不同欄位的寫入配置
		self.current_field = None
		self.empty_cells = []  # 當前欄位的空格

		# 初始化界面變量
		self.config_var = tk.StringVar()
		self.first_keyword_var = tk.StringVar()  # 第一個關鍵字
		self.field_var = tk.StringVar()  # 第二個關鍵字
		self.new_config_var = tk.StringVar()

		# 建立寫入資料夾
		os.makedirs(FIELD_MAPPING_DIR, exist_ok=True)

		self.setup_ui()
		self.load_configs()

		# 啟動時自動嘗試連接Excel
		self.root.after(500, self.auto_connect_excel)

		# 定期檢查Excel連接狀態
		self.start_excel_monitoring()

		# 初始化彈出式選單變數
		self.config_popup = None

	def setup_ui(self):
		"""設置界面"""
		# 設置自定義樣式
		style = ttk.Style()
		style.configure("Large.TButton", font=('Arial', 10), padding=6)
		style.configure("Large.TLabel", font=('Arial', 10))
		style.configure("Large.TRadiobutton", font=('Arial', 10))

		# 主框架
		main_frame = ttk.Frame(self.root)
		main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

		# =================== 簡潔控制面板 ===================
		control_frame = ttk.LabelFrame(main_frame, text="控制面板", padding=10)
		control_frame.pack(fill=tk.X, pady=(0, 10))

		# 第一行：檔案和Excel
		row1 = ttk.Frame(control_frame)
		row1.pack(fill=tk.X, pady=(0, 10))

		# 檔案操作群組
		file_group = ttk.Frame(row1)
		file_group.pack(side=tk.LEFT, padx=(0, 30))

		self.csv_btn = ttk.Button(file_group, text="載入CSV", command=self.load_csv, width=12,
									style="Large.TButton")
		self.csv_btn.pack(side=tk.LEFT, padx=(0, 10))

		self.manual_connect_btn = ttk.Button(file_group, text="連接Excel", command=self.connect_excel, width=12,
											style="Large.TButton")
		# 預設不顯示，由toggle_connection_mode控制

		# Excel模式和狀態群組
		excel_group = ttk.Frame(row1)
		excel_group.pack(side=tk.LEFT, padx=(0, 20))

		# Excel模式選擇
		mode_frame = ttk.Frame(excel_group)
		mode_frame.pack(side=tk.TOP, anchor=tk.W)

		self.mode_var = tk.BooleanVar(value=True)
		self.auto_radio = ttk.Radiobutton(mode_frame, text="自動偵測", variable=self.mode_var,
											value=True, command=self.toggle_connection_mode,
											style="Large.TRadiobutton")
		self.auto_radio.pack(side=tk.LEFT, padx=(0, 15))

		self.manual_radio = ttk.Radiobutton(mode_frame, text="手動連接", variable=self.mode_var,
													value=False, command=self.toggle_connection_mode,
													style="Large.TRadiobutton")
		self.manual_radio.pack(side=tk.LEFT)

		# 隱藏的Excel狀態（僅供內部使用）
		self.excel_status = ttk.Label(mode_frame, text="正在嘗試連接Excel...", foreground="orange")
		# 不要pack，保持隱藏狀態

		# 寫入區域（右側）
		execute_main_group = ttk.Frame(row1)
		execute_main_group.pack(side=tk.RIGHT)

		# 檔案名稱顯示區域（加長版，緊貼寫入左側）
		file_info_group = ttk.LabelFrame(execute_main_group, text="當前檔案", padding=5)
		file_info_group.pack(side=tk.LEFT, padx=(0, 10))

		# 固定寬度的框架
		file_info_frame = ttk.Frame(file_info_group)
		file_info_frame.pack()

		# CSV檔案名稱顯示（加長）
		csv_name_frame = ttk.Frame(file_info_frame)
		csv_name_frame.pack(pady=(0, 5))
		ttk.Label(csv_name_frame, text="CSV  :", font=('Arial', 10, 'bold')).pack(side=tk.LEFT)
		self.csv_name_label = ttk.Label(csv_name_frame, text="未載入",
												font=('Arial', 10), foreground="gray", width=35)
		self.csv_name_label.pack(side=tk.LEFT, padx=(8, 0))

		# Excel檔案名稱顯示（加長）
		excel_name_frame = ttk.Frame(file_info_frame)
		excel_name_frame.pack()
		ttk.Label(excel_name_frame, text="Excel:", font=('Arial', 10, 'bold')).pack(side=tk.LEFT)
		self.excel_name_label = ttk.Label(excel_name_frame, text="未連接",
											font=('Arial', 10), foreground="gray", width=35)
		self.excel_name_label.pack(side=tk.LEFT, padx=(8, 0))

		# 創建寫入按鈕（大的方形，醒目顏色）
		self.execute_btn = tk.Button(execute_main_group, text="寫入",
									command=self.execute_smart_mapping,
									width=15, height=3,
									background="lightblue", foreground="black",
									font=('Arial', 12, 'bold'),
									relief=tk.RAISED, bd=3,
									activebackground="lightgreen",
									cursor="hand2")
		self.execute_btn.pack(side=tk.LEFT, padx=5, pady=5)

		# 第二行：配置區
		config_frame = ttk.LabelFrame(control_frame, text="配置管理區", padding=8)
		config_frame.pack(fill=tk.X, pady=(0, 10))

		config_content = ttk.Frame(config_frame)
		config_content.pack(fill=tk.X)

		# 當前配置選擇
		current_config_group = ttk.Frame(config_content)
		current_config_group.pack(side=tk.LEFT, padx=(0, 30))

		ttk.Label(current_config_group, text="當前配置:", style="Large.TLabel").pack(side=tk.LEFT, padx=(0, 5))

		config_click_frame = ttk.Frame(current_config_group)
		config_click_frame.pack(side=tk.LEFT, padx=(0, 10))

		self.config_display = ttk.Entry(config_click_frame, textvariable=self.config_var, state='readonly',
										width=25, font=('Arial', 10))
		self.config_display.pack(side=tk.LEFT)

		# 隱藏的Combobox
		self.config_combo = ttk.Combobox(control_frame, textvariable=self.config_var, width=1)
		self.config_combo.place(x=-1000, y=-1000)

		# 綁定事件
		config_click_frame.bind('<Button-1>', self.on_config_click)
		self.config_display.bind('<Button-1>', self.on_config_click)

		# 刪除按鈕
		ttk.Button(current_config_group, text="刪除配置", command=self.delete_config, width=12,
							style="Large.TButton").pack(side=tk.LEFT)

		# 新配置保存
		new_config_group = ttk.Frame(config_content)
		new_config_group.pack(side=tk.LEFT)

		ttk.Label(new_config_group, text="新配置名稱:", style="Large.TLabel").pack(side=tk.LEFT, padx=(0, 5))
		ttk.Entry(new_config_group, textvariable=self.new_config_var, width=25, font=('Arial', 10)).pack(side=tk.LEFT, padx=(0, 10))
		ttk.Button(new_config_group, text="保存配置", command=self.save_config, width=12,
							style="Large.TButton").pack(side=tk.LEFT)

		# 第三行：關鍵字輸入區
		keyword_frame = ttk.LabelFrame(control_frame, text="關鍵字設定區", padding=8)
		keyword_frame.pack(fill=tk.X, pady=(0, 10))

		keyword_content = ttk.Frame(keyword_frame)
		keyword_content.pack(fill=tk.X)

		# 第一關鍵字
		first_keyword_group = ttk.Frame(keyword_content)
		first_keyword_group.pack(side=tk.LEFT, padx=(0, 30))

		ttk.Label(first_keyword_group, text="第一關鍵字:", style="Large.TLabel").pack(side=tk.LEFT, padx=(0, 5))
		ttk.Entry(first_keyword_group, textvariable=self.first_keyword_var, width=25, font=('Arial', 10)).pack(side=tk.LEFT)

		# 第二關鍵字
		second_keyword_group = ttk.Frame(keyword_content)
		second_keyword_group.pack(side=tk.LEFT)

		ttk.Label(second_keyword_group, text="第二關鍵字:", style="Large.TLabel").pack(side=tk.LEFT, padx=(0, 5))
		ttk.Entry(second_keyword_group, textvariable=self.field_var, width=25, font=('Arial', 10)).pack(side=tk.LEFT)

		# 第四行：獲取空格位置使用選取範圍區
		scan_frame = ttk.LabelFrame(control_frame, text="獲取空格位置 & 手動選取儲存格區", padding=8)
		scan_frame.pack(fill=tk.X, pady=(0, 5))

		scan_content = ttk.Frame(scan_frame)
		scan_content.pack(fill=tk.X)

		# 掃描操作按鈕
		scan_buttons_group = ttk.Frame(scan_content)
		scan_buttons_group.pack(side=tk.LEFT)

		ttk.Button(scan_buttons_group, text="獲取空格位置", command=self.scan_empty_cells, width=15,
							style="Large.TButton").pack(side=tk.LEFT, padx=(0, 10))

		ttk.Button(scan_buttons_group, text="手動選取儲存格", command=self.scan_selection_range, width=15,
							style="Large.TButton").pack(side=tk.LEFT)

		# =================== 主工作區 ===================
		work_frame = ttk.Frame(main_frame)
		work_frame.pack(fill=tk.BOTH, expand=True)

		# 左側：CSV選擇區
		csv_frame = ttk.LabelFrame(work_frame, text="CSV數據 (選擇Elements)", padding=5)
		csv_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

		# CSV表格
		csv_table_frame = ttk.Frame(csv_frame)
		csv_table_frame.pack(fill=tk.BOTH, expand=True)

		# 建立Treeview表格
		self.csv_tree = ttk.Treeview(csv_table_frame, selectmode="extended", height=25)
		csv_scroll_y = ttk.Scrollbar(csv_table_frame, orient=tk.VERTICAL, command=self.csv_tree.yview)
		self.csv_tree.configure(yscrollcommand=csv_scroll_y.set)

		# 綁定點擊事件，實現單擊切換選取狀態
		self.csv_tree.bind('<Button-1>', self.on_tree_click)

		# 設定表格欄位
		self.csv_tree['columns'] = ('Element', 'Dev', 'Actual', 'Value')
		self.csv_tree['show'] = 'headings'

		# 設定欄位標題和寬度
		self.csv_tree.heading('Element', text='Element')
		self.csv_tree.heading('Dev', text='Dev')
		self.csv_tree.heading('Actual', text='Actual')
		self.csv_tree.heading('Value', text='使用值')

		self.csv_tree.column('Element', width=80, minwidth=60)
		self.csv_tree.column('Dev', width=80, minwidth=60, anchor='center')
		self.csv_tree.column('Actual', width=80, minwidth=60, anchor='center')
		self.csv_tree.column('Value', width=80, minwidth=60, anchor='center')

		self.csv_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
		csv_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

		# 選取數量顯示和操作按鈕
		csv_info_frame = ttk.Frame(csv_frame)
		csv_info_frame.pack(fill=tk.X, pady=(5, 0))

		# 左側：選取數量和說明
		csv_info_left = ttk.Frame(csv_info_frame)
		csv_info_left.pack(side=tk.LEFT, fill=tk.X, expand=True)

		self.csv_selection_label = ttk.Label(csv_info_left, text="已選取: 0 個元素", font=('Arial', 11, 'bold'))
		self.csv_selection_label.pack(anchor=tk.W)

		# 使用說明
		csv_help_label = ttk.Label(csv_info_left, text="操作說明: 點擊選取項目，再次點擊取消選取", font=('Arial', 9), foreground="gray")
		csv_help_label.pack(anchor=tk.W)

		# 右側：操作按鈕
		csv_buttons_frame = ttk.Frame(csv_info_frame)
		csv_buttons_frame.pack(side=tk.RIGHT, padx=(10, 0))

		ttk.Button(csv_buttons_frame, text="全選", command=self.select_all_csv, width=10,
					style="Large.TButton").pack(side=tk.LEFT, padx=(0, 5))

		ttk.Button(csv_buttons_frame, text="取消全選", command=self.deselect_all_csv, width=10,
					style="Large.TButton").pack(side=tk.LEFT)

		# 綁定選取事件
		self.csv_tree.bind('<<TreeviewSelect>>', self.update_selection_info)

		# 右側：寫入配置區
		mapping_frame = ttk.LabelFrame(work_frame, text="寫入配置與狀態", padding=5)
		mapping_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))

		# 狀態顯示區
		status_frame = ttk.LabelFrame(mapping_frame, text="寫入狀態", padding=5)
		status_frame.pack(fill=tk.X, pady=(10, 10))

		# 空格數量顯示
		self.spaces_count_label = ttk.Label(status_frame, text="找到空格: 0 個", font=('Arial', 11, 'bold'))
		self.spaces_count_label.pack(anchor=tk.W, pady=(0, 5))

		# 匹配狀態顯示
		self.match_status_label = ttk.Label(status_frame, text="數量不匹配", foreground="orange", font=('Arial', 11, 'bold'))
		self.match_status_label.pack(anchor=tk.W)

		# 空格信息詳細顯示
		info_frame = ttk.LabelFrame(mapping_frame, text="空格位置詳細", padding=5)
		info_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

		self.empty_cells_info = tk.Text(info_frame, height=15, wrap=tk.WORD, font=('Consolas', 9))
		info_scroll = ttk.Scrollbar(info_frame, orient=tk.VERTICAL, command=self.empty_cells_info.yview)
		self.empty_cells_info.configure(yscrollcommand=info_scroll.set)

		self.empty_cells_info.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
		info_scroll.pack(side=tk.RIGHT, fill=tk.Y)


	def on_config_click(self, event):
		"""處理配置選項點擊事件"""
		# 如果彈出視窗已存在，先銷毀它
		if self.config_popup:
			self.config_popup.destroy()
			self.config_popup = None
			return

		# 創建彈出視窗
		self.config_popup = tk.Toplevel(self.root)
		self.config_popup.title("選擇配置")
		self.config_popup.geometry("300x200")
		self.config_popup.resizable(False, False)

		# 設定彈出視窗位置（在點擊位置附近）
		x = self.root.winfo_rootx() + event.x_root - self.root.winfo_rootx()
		y = self.root.winfo_rooty() + event.y_root - self.root.winfo_rooty() + 30
		self.config_popup.geometry(f"+{x}+{y}")

		# 創建列表框
		listbox_frame = ttk.Frame(self.config_popup)
		listbox_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

		# 配置列表
		config_listbox = tk.Listbox(listbox_frame, height=8)
		config_scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=config_listbox.yview)
		config_listbox.configure(yscrollcommand=config_scrollbar.set)

		# 添加配置選項
		config_names = list(self.field_mappings.keys())
		for name in config_names:
			config_listbox.insert(tk.END, name)

		# 選中當前配置
		current_config = self.config_var.get()
		if current_config in config_names:
			config_listbox.selection_set(config_names.index(current_config))

		config_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
		config_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

		# 按鈕框架
		button_frame = ttk.Frame(self.config_popup)
		button_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

		def on_select():
			selection = config_listbox.curselection()
			if selection:
				selected_config = config_listbox.get(selection[0])
				self.config_var.set(selected_config)
				# 自動套用選中的配置
				self.load_config()
			self.config_popup.destroy()
			self.config_popup = None

		def on_cancel():
			self.config_popup.destroy()
			self.config_popup = None

		ttk.Button(button_frame, text="取消", command=on_cancel).pack(side=tk.RIGHT)
		ttk.Button(button_frame, text="確定", command=on_select).pack(side=tk.RIGHT)

		# 綁定雙擊事件（自動套用配置）
		config_listbox.bind('<Double-Button-1>', lambda e: on_select())

		# 綁定ESC鍵關閉
		self.config_popup.bind('<Escape>', lambda e: on_cancel())

		# 設定焦點
		self.config_popup.focus_set()
		config_listbox.focus_set()

		# 讓彈出視窗保持在最上層
		self.config_popup.transient(self.root)
		self.config_popup.grab_set()

	def load_csv(self):
		"""載入CSV文件"""
		file_path = filedialog.askopenfilename(
			title="選擇CSV文件",
			filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
		)

		if file_path:
			try:
				with open(file_path, 'r', encoding='utf-8') as f:
					reader = csv.DictReader(f)
					self.csv_data = list(reader)

				# 更新CSV檔案名稱顯示
				csv_filename = os.path.basename(file_path)
				self.csv_name_label.config(text=csv_filename, foreground="black")

				self.display_csv_data()

				# 自動套用當前選中的配置
				self.auto_apply_current_config()

			except Exception as e:
				messagebox.showerror("錯誤", f"載入CSV失敗：{str(e)}")

	def display_csv_data(self):
		"""顯示CSV數據"""
		# 清除舊數據
		for item in self.csv_tree.get_children():
			self.csv_tree.delete(item)

		if not self.csv_data:
			return

		for i, row in enumerate(self.csv_data):
			element = row.get('Element', f'行{i+1}')
			dev = row.get('Dev', '')
			actual = row.get('Actual', '')

			# 決定使用哪個值（Dev優先，沒有則用Actual）
			use_value = ""
			if dev and str(dev).strip() and str(dev).strip().lower() != 'n/a':
				use_value = str(dev).strip()
			elif actual and str(actual).strip() and str(actual).strip().lower() != 'n/a':
				use_value = str(actual).strip()

			# 顯示格式化的值
			dev_display = self.get_display_value(dev)
			actual_display = self.get_display_value(actual)
			value_display = use_value if use_value else '-'

			self.csv_tree.insert('', 'end', values=(element, dev_display, actual_display, value_display))

	def on_tree_click(self, event):
		"""處理Treeview點擊事件，實現單擊切換選取"""
		# 獲取點擊的項目
		item = self.csv_tree.identify_row(event.y)

		if item:
			# 獲取目前選取的項目
			current_selection = self.csv_tree.selection()

			# 如果項目已被選取，則取消選取；否則加入選取
			if item in current_selection:
				# 取消選取這個項目
				new_selection = [i for i in current_selection if i != item]
			else:
				# 加入選取
				new_selection = list(current_selection) + [item]

			# 更新選取狀態
			self.csv_tree.selection_set(new_selection)

			# 阻止預設的選取行為
			return "break"

		# 延遲更新顯示，確保選取狀態已更新
		self.csv_tree.after(10, self.update_selection_info)

	def update_selection_info(self, event=None):
		"""更新選取信息"""
		selected_count = len(self.csv_tree.selection())
		self.csv_selection_label.config(text=f"已選取: {selected_count} 個元素")

		# 更新匹配狀態
		self.update_match_status()

	def select_all_csv(self):
		"""全選CSV中的所有元素"""
		all_items = self.csv_tree.get_children()
		self.csv_tree.selection_set(all_items)
		self.update_selection_info()

	def deselect_all_csv(self):
		"""取消全選CSV中的所有元素"""
		self.csv_tree.selection_remove(self.csv_tree.selection())
		self.update_selection_info()

	def update_match_status(self):
		"""更新匹配狀態顯示"""
		selected_count = len(self.csv_tree.selection())
		spaces_count = len(self.empty_cells)

		if spaces_count == 0:
			self.match_status_label.config(text="請先獲取空格位置", foreground="orange")
		elif selected_count == spaces_count and selected_count > 0:
			self.match_status_label.config(text="數量匹配，可以執行", foreground="green")
		elif selected_count == 0:
			self.match_status_label.config(text="請選擇CSV元素", foreground="orange")
		else:
			self.match_status_label.config(text=f"數量不匹配 ({selected_count}/{spaces_count})", foreground="red")

	def update_excel_name_display(self, name, color="black"):
		"""統一更新Excel檔案名稱顯示"""
		self.excel_name_label.config(text=name, foreground=color)

	def is_cell_empty(self, cell_value):
		"""檢查儲存格是否為空"""
		return cell_value is None or str(cell_value).strip() == ''

	def get_display_value(self, value):
		"""獲取顯示用的值"""
		return str(value) if value and str(value).strip() else '-'

	def get_excel_column_name(self, col_index):
		"""將數字索引轉換為Excel列名（A, B, ..., Z, AA, AB, ...）"""
		column_name = ""
		while col_index >= 0:
			column_name = chr(65 + (col_index % 26)) + column_name
			col_index = col_index // 26 - 1
		return column_name

	def auto_connect_excel(self):
		"""啟動時自動連接Excel（靜默模式）"""
		try:
			excel_app = win32com.client.GetActiveObject("Excel.Application")
			self.active_workbook = excel_app.ActiveWorkbook

			if self.active_workbook:
				# 使用當前活動的工作表
				self.active_worksheet = self.active_workbook.ActiveSheet

				workbook_name = self.active_workbook.Name
				self.excel_status.config(text=f"已連接: {workbook_name}", foreground="green")
				self.update_excel_name_display(workbook_name)
				self.load_excel_data()
			else:
				raise Exception("沒有開啟的工作簿")

		except:
			# 靜默失敗，在自動模式下顯示等待狀態
			if self.auto_detect_mode:
				self.update_excel_name_display("等待中", "gray")
			else:
				self.update_excel_name_display("未連接", "gray")

	def start_excel_monitoring(self):
		"""開始監控Excel狀態"""
		self.monitor_excel()

	def monitor_excel(self):
		"""定期監控Excel連接狀態（僅在自動偵測模式下）"""
		# 只在自動偵測模式下執行監控
		if not self.auto_detect_mode:
			# 如果不是自動偵測模式，3秒後再檢查
			self.root.after(3000, self.monitor_excel)
			return

		try:
			current_status = self.check_excel_status()

			# 獲取當前顯示的狀態
			current_display = self.excel_status.cget("text")

			# 如果狀態發生變化，更新顯示
			if current_status != current_display:
				if current_status.startswith("已連接"):
					# Excel重新連接，嘗試載入數據
					if not current_display.startswith("已連接"):
						self.load_excel_data()
						# 如果有配置，自動重新獲取空格位置
						self.auto_rescan_on_reconnect()
						print(f"Excel狀態變化: {current_display} → {current_status}")

				# 根據狀態設定顏色
				if current_status.startswith("已連接"):
					color = "green"
				elif current_status.startswith("等待"):
					color = "orange"
				else:
					color = "red"

				self.excel_status.config(text=current_status, foreground=color)

		except Exception as e:
			# 監控過程中的錯誤不要打擾用戶
			pass

		# 每3秒檢查一次
		self.root.after(3000, self.monitor_excel)

	def check_excel_status(self):
		"""檢查Excel當前狀態"""
		try:
			excel_app = win32com.client.GetActiveObject("Excel.Application")
			active_workbook = excel_app.ActiveWorkbook

			if active_workbook:
				workbook_name = active_workbook.Name
				# 更新實例變數，使用當前活動的工作表
				self.active_workbook = active_workbook
				self.active_worksheet = active_workbook.ActiveSheet

				self.update_excel_name_display(workbook_name)
				return f"已連接: {workbook_name}"
			else:
				# Excel開啟但沒有工作簿
				self.active_workbook = None
				self.active_worksheet = None
				self.update_excel_name_display("無工作簿", "gray")
				return "Excel已開啟但無工作簿"

		except:
			# Excel未開啟或連接失敗
			self.active_workbook = None
			self.active_worksheet = None
			self.update_excel_name_display("未連接", "gray")
			# 在自動模式下顯示等待，手動模式顯示未連接
			if self.auto_detect_mode:
				return "等待Excel開啟..."
			else:
				return "未連接Excel"

		# 在自動模式下顯示等待，手動模式顯示未連接
		if self.auto_detect_mode:
			self.update_excel_name_display("等待中", "gray")
			return "等待Excel開啟..."
		else:
			self.update_excel_name_display("未連接", "gray")
			return "未連接Excel"

	def auto_rescan_on_reconnect(self):
		"""Excel重新連接時自動重新掃描"""
		try:
			# 檢查是否有選中的配置和目標欄位
			current_config = self.config_var.get().strip()
			target_field = self.field_var.get().strip()

			if current_config and target_field:
				# 檢查是否為選取範圍模式
				config_data = self.field_mappings.get(current_config, {})
				use_selection_mode = config_data.get('use_selection_mode', False)

				# 只有在非選取範圍模式下才自動掃描
				if not use_selection_mode:
					# 靜默重新獲取空格位置
					self.scan_empty_cells()
					print(f"Excel重新連接，已自動重新掃描欄位: {target_field}")

		except Exception as e:
			# 靜默處理錯誤
			print(f"自動重新掃描失敗: {e}")

	def toggle_connection_mode(self):
		"""切換連接模式"""
		self.auto_detect_mode = self.mode_var.get()

		if self.auto_detect_mode:
			# 切換到自動偵測模式
			self.manual_connect_btn.pack_forget()  # 隱藏手動連接按鈕
			# 立即嘗試自動連接
			self.root.after(100, self.auto_connect_excel)
		else:
			# 切換到手動選取模式 - 在檔案群組中顯示連接按鈕
			self.manual_connect_btn.pack(side=tk.LEFT)
			self.update_excel_name_display("未連接", "gray")

	def connect_excel(self):
		"""手動連接Excel"""
		self.connect_excel_windows()

	def connect_excel_windows(self):
		"""Windows連接"""
		try:
			excel_app = win32com.client.GetActiveObject("Excel.Application")
			self.active_workbook = excel_app.ActiveWorkbook

			if not self.active_workbook:
				raise Exception("沒有開啟的工作簿")

			# 使用當前活動的工作表
			self.active_worksheet = self.active_workbook.ActiveSheet

			workbook_name = self.active_workbook.Name
			self.excel_status.config(text=f"已連接: {workbook_name}", foreground="green")
			self.update_excel_name_display(workbook_name)
			self.load_excel_data()

		except Exception as e:
			# 在自動模式下，不做任何操作，讓監控繼續等待Excel開啟
			# 在手動模式下，直接跳轉到文件選擇
			if not self.auto_detect_mode:
				self.manual_excel_setup()

	def manual_excel_setup(self):
		"""手動選擇Excel文件"""
		# 在手動模式下，直接選擇文件而不顯示"連接失敗"訊息
		if not self.auto_detect_mode:
			result = True
		else:
			result = messagebox.askyesno("連接失敗",
				"無法自動連接到Excel。是否選擇Excel文件？")

		if result:
			file_path = filedialog.askopenfilename(
				title="選擇Excel文件",
				filetypes=[("Excel files", "*.xlsx"), ("Excel files", "*.xls")]
			)

			if file_path:
				try:
					self.excel_workbook = load_workbook(file_path, data_only=True)
					# 使用活動的工作表
					self.excel_sheet = self.excel_workbook.active
					filename = os.path.basename(file_path)
					self.excel_status.config(text=f"已載入: {filename}", foreground="blue")
					self.update_excel_name_display(filename, "blue")
					self.load_excel_data()
				except Exception as e:
					messagebox.showerror("錯誤", f"載入Excel失敗：{str(e)}")

	def load_excel_data(self):
		"""載入Excel數據"""
		try:
			if self.active_worksheet:
				# 從COM接口讀取
				used_range = self.active_worksheet.UsedRange
				values = used_range.Value
				if isinstance(values, tuple):
					self.excel_data = [list(row) if isinstance(row, tuple) else [row] for row in values]
				else:
					self.excel_data = [[values]]
			elif self.excel_sheet:
				# 從openpyxl讀取
				self.excel_data = []
				for row in self.excel_sheet.iter_rows(values_only=True):
					self.excel_data.append(list(row))

		except Exception as e:
			messagebox.showerror("錯誤", f"載入Excel數據失敗：{str(e)}")

	def find_field_position(self, field_name):
		"""尋找欄位位置"""
		for row_idx, row in enumerate(self.excel_data):
			for col_idx, cell in enumerate(row):
				if cell and field_name in str(cell):
					return (row_idx, col_idx)
		return None

	def find_second_keyword_in_column(self, first_row, col_idx, second_keyword):
		"""在同一列中往下尋找第二個關鍵字"""
		# 從第一個關鍵字的下一行開始往下找
		for row_idx in range(first_row + 1, len(self.excel_data)):
			if col_idx < len(self.excel_data[row_idx]):
				cell = self.excel_data[row_idx][col_idx]
				if cell and second_keyword in str(cell):
					return row_idx
		return None

	def scan_vertical_empty_cells(self, field_row, field_col):
		"""垂直獲取空格位置"""
		empty_cells = []
		current_row = field_row + 1

		while current_row < len(self.excel_data):
			row_empty_count = 0
			row_has_content = False

			for col_offset in range(MAX_HORIZONTAL_SCAN_RANGE):
				check_col = field_col + col_offset
				if check_col < len(self.excel_data[current_row]):
					cell_value = self.excel_data[current_row][check_col]
					cell_str = str(cell_value).strip() if cell_value is not None else ''

					if not self.is_cell_empty(cell_value):
						row_has_content = True
						break

					col_letter = self.get_excel_column_name(check_col)
					cell_position = f"{col_letter}{current_row + 1}"
					empty_cells.append({
						'position': cell_position,
						'row': current_row,
						'col': check_col,
						'value': cell_value
					})
					row_empty_count += 1
				else:
					break

			if row_has_content or row_empty_count == 0:
				break

			current_row += 1

		return empty_cells

	def scan_horizontal_empty_cells(self, field_row, field_col):
		"""水平獲取空格位置"""
		empty_cells = []
		for col_offset in range(1, MAX_VERTICAL_SCAN_RANGE):
			check_col = field_col + col_offset
			check_row = field_row + 1
			if (check_row < len(self.excel_data) and
				check_col < len(self.excel_data[check_row])):
				cell_value = self.excel_data[check_row][check_col]
				if self.is_cell_empty(cell_value):
					col_letter = self.get_excel_column_name(check_col)
					cell_position = f"{col_letter}{check_row + 1}"
					empty_cells.append({
						'position': cell_position,
						'row': check_row,
						'col': check_col,
						'value': cell_value
					})
		return empty_cells

	def scan_empty_cells(self):
		"""獲取空格位置（支援兩段定位）"""
		first_keyword = self.first_keyword_var.get().strip()
		second_keyword = self.field_var.get().strip()

		if not second_keyword:
			messagebox.showwarning("警告", "請輸入第二關鍵字")
			return

		if not self.excel_data:
			messagebox.showwarning("警告", "請先連接Excel")
			return

		try:
			field_row = None
			field_col = None

			# 如果有第一個關鍵字，使用兩段定位
			if first_keyword:
				# 1. 先找第一個關鍵字
				first_position = self.find_field_position(first_keyword)
				if not first_position:
					messagebox.showwarning("警告", f"找不到第一關鍵字: {first_keyword}")
					return

				first_row, first_col = first_position

				# 2. 在同一列往下找第二個關鍵字
				second_row = self.find_second_keyword_in_column(first_row, first_col, second_keyword)
				if second_row is None:
					messagebox.showwarning("警告",
						f"在第一關鍵字 '{first_keyword}' 的同一列中\n往下找不到第二關鍵字: {second_keyword}")
					return

				field_row = second_row
				field_col = first_col

			else:
				# 沒有第一個關鍵字，使用原來的邏輯（直接找第二個關鍵字）
				field_position = self.find_field_position(second_keyword)
				if not field_position:
					messagebox.showwarning("警告", f"找不到第二關鍵字: {second_keyword}")
					return

				field_row, field_col = field_position

			# 3. 在第二個關鍵字下方獲取空格位置
			empty_cells = self.scan_vertical_empty_cells(field_row, field_col)

			if not empty_cells:
				empty_cells = self.scan_horizontal_empty_cells(field_row, field_col)

			self.empty_cells = empty_cells
			self.display_empty_cells_info()
			self.spaces_count_label.config(text=f"找到空格: {len(empty_cells)} 個")
			self.update_match_status()

			if not empty_cells:
				if first_keyword:
					messagebox.showwarning("警告",
						f"在第二關鍵字 '{second_keyword}' 下方沒有找到空白位置")
				else:
					messagebox.showwarning("警告",
						f"在關鍵字 '{second_keyword}' 下方沒有找到空白位置")

		except Exception as e:
			messagebox.showerror("錯誤", f"掃描失敗：{str(e)}")

	def scan_selection_range(self):
		"""使用Excel中的選取範圍作為目標位置"""
		if not self.excel_data:
			messagebox.showwarning("警告", "請先連接Excel")
			return

		try:
			# 檢查是否有活動的工作表
			if not self.active_worksheet:
				messagebox.showwarning("警告", "此功能需要COM模式連接Excel\n請確認Excel已開啟並處於自動偵測模式")
				return

			# 獲取選取範圍
			selection = self.active_worksheet.Application.Selection

			empty_cells = []

			# 處理選取範圍
			# 處理單個儲存格或範圍
			try:
				# 嘗試遍歷選取的儲存格
				for cell in selection:
					row_num = cell.Row - 1  # 轉換為0-based索引
					col_num = cell.Column - 1  # 轉換為0-based索引
					col_letter = self.get_excel_column_name(col_num)
					cell_position = f"{col_letter}{row_num + 1}"

					empty_cells.append({
						'position': cell_position,
						'row': row_num,
						'col': col_num,
						'value': cell.Value
					})
			except:
				# 如果是單個儲存格
				row_num = selection.Row - 1
				col_num = selection.Column - 1
				col_letter = self.get_excel_column_name(col_num)
				cell_position = f"{col_letter}{row_num + 1}"

				empty_cells.append({
					'position': cell_position,
					'row': row_num,
					'col': col_num,
					'value': selection.Value
				})

			if not empty_cells:
				messagebox.showwarning("警告", "未選取任何儲存格")
				return

			# 更新空格列表
			self.empty_cells = empty_cells
			self.display_empty_cells_info_for_selection()
			self.spaces_count_label.config(text=f"選取範圍: {len(empty_cells)} 個儲存格")
			self.update_match_status()

		except Exception as e:
			messagebox.showerror("錯誤", f"讀取選取範圍失敗：{str(e)}")

	def display_empty_cells_info(self):
		"""顯示空格信息"""
		self.empty_cells_info.delete(1.0, tk.END)

		if not self.empty_cells:
			self.empty_cells_info.insert(tk.END, "尚未掃描到可填入位置\n\n請點擊'獲取空格位置'")
			return

		first_keyword = self.first_keyword_var.get().strip()
		second_keyword = self.field_var.get().strip()

		if first_keyword:
			self.empty_cells_info.insert(tk.END, f"第一關鍵字: {first_keyword}\n")
			self.empty_cells_info.insert(tk.END, f"第二關鍵字: {second_keyword}\n")
		else:
			self.empty_cells_info.insert(tk.END, f"關鍵字: {second_keyword}\n")

		self.empty_cells_info.insert(tk.END, f"找到 {len(self.empty_cells)} 個位置:\n\n")

		for i, cell in enumerate(self.empty_cells):
			self.empty_cells_info.insert(tk.END, f"{i+1}. {cell['position']}\n")

		self.empty_cells_info.insert(tk.END, f"\n請在左側選擇 {len(self.empty_cells)} 個CSV元素")

	def display_empty_cells_info_for_selection(self):
		"""顯示選取範圍信息"""
		self.empty_cells_info.delete(1.0, tk.END)

		if not self.empty_cells:
			return

		self.empty_cells_info.insert(tk.END, "使用Excel選取範圍\n")
		self.empty_cells_info.insert(tk.END, f"找到 {len(self.empty_cells)} 個儲存格:\n\n")

		for i, cell in enumerate(self.empty_cells):
			self.empty_cells_info.insert(tk.END, f"{i+1}. {cell['position']}\n")

		self.empty_cells_info.insert(tk.END, f"\n請在左側選擇 {len(self.empty_cells)} 個CSV元素")

	def execute_smart_mapping(self):
		"""寫入"""
		# 獲取選中的CSV項目
		selected_items = self.csv_tree.selection()

		# 詳細的防呆檢查
		if not self.csv_data:
			messagebox.showerror("錯誤", "請先載入CSV文件")
			return

		if not selected_items:
			messagebox.showwarning("警告", "請在左側選擇CSV中的元素")
			return

		if not self.empty_cells:
			messagebox.showwarning("警告", "請先點擊'獲取空格位置'找到可填入的位置")
			return

		if len(selected_items) != len(self.empty_cells):
			messagebox.showerror("錯誤",
				f"數量必須完全匹配！\n\n"
				f"您選了: {len(selected_items)} 個CSV元素\n"
				f"找到: {len(self.empty_cells)} 個空格位置\n\n"
				f"請重新選擇，確保數量一致")
			return

		# 檢查Excel連接狀態
		if not self.active_worksheet and not self.excel_sheet:
			messagebox.showerror("錯誤", "Excel連接已斷開，請重新連接Excel")
			return

		# 確認執行
		first_keyword = self.first_keyword_var.get().strip()
		second_keyword = self.field_var.get().strip()

		if first_keyword:
			confirm_msg = (f"即將填入 {len(selected_items)} 個數據到Excel\n\n"
				f"第一關鍵字: {first_keyword}\n"
				f"第二關鍵字: {second_keyword}\n"
				f"確定要執行嗎？")
		else:
			confirm_msg = (f"即將填入 {len(selected_items)} 個數據到Excel\n\n"
				f"目標關鍵字: {second_keyword}\n"
				f"確定要執行嗎？")

		result = messagebox.askyesno("確認執行", confirm_msg)

		if not result:
			return

		try:
			filled_count = 0

			for i, item_id in enumerate(selected_items):
				# 獲取選中項目的索引
				item_index = self.csv_tree.index(item_id)
				csv_row = self.csv_data[item_index]

				# 決定使用Dev還是Actual（Dev優先）
				dev_value = csv_row.get('Dev', '')
				actual_value = csv_row.get('Actual', '')

				use_value = None
				if dev_value and str(dev_value).strip() and str(dev_value).strip().lower() != 'n/a':
					use_value = str(dev_value).strip()
				elif actual_value and str(actual_value).strip() and str(actual_value).strip().lower() != 'n/a':
					use_value = str(actual_value).strip()

				if use_value:
					empty_cell = self.empty_cells[i]
					row_num = empty_cell['row'] + 1  # Excel行從1開始
					col_num = empty_cell['col'] + 1  # Excel列從1開始

					# 轉換數值
					try:
						numeric_value = float(use_value)
					except ValueError:
						numeric_value = use_value

					# 填入數據
					if self.active_worksheet:
						self.active_worksheet.Cells(row_num, col_num).Value = numeric_value
					elif self.excel_sheet:
						self.excel_sheet.cell(row=row_num, column=col_num, value=numeric_value)

					filled_count += 1

			if self.excel_workbook and not self.active_worksheet:
				# openpyxl模式，自動儲存（覆寫原檔案）
				try:
					if hasattr(self.excel_workbook, 'filename') and self.excel_workbook.filename:
						self.excel_workbook.save(self.excel_workbook.filename)
					else:
						# 若無原始檔名，則另存新檔
						save_path = filedialog.asksaveasfilename(
							title="保存Excel文件",
							defaultextension=".xlsx",
							filetypes=[("Excel files", "*.xlsx")]
						)
						if save_path:
							self.excel_workbook.save(save_path)
				except Exception as e:
					messagebox.showwarning("警告", f"自動儲存Excel失敗：{str(e)}")
			elif self.active_workbook:
				# Windows COM模式，自動儲存
				try:
					self.active_workbook.Save()
				except Exception as e:
					messagebox.showwarning("警告", f"自動儲存Excel失敗：{str(e)}")

			# 構建成功訊息
			if first_keyword:
				success_msg = (f"寫入完成！\n\n"
					f"已成功填入 {filled_count} 個數據\n"
					f"第一關鍵字: {first_keyword}\n"
					f"第二關鍵字: {second_keyword}\n"
					f"請檢查Excel文件確認結果")
			else:
				success_msg = (f"寫入完成！\n\n"
					f"已成功填入 {filled_count} 個數據到關鍵字: {second_keyword}\n"
					f"請檢查Excel文件確認結果")

			messagebox.showinfo("成功", success_msg)

			# 清空CSV資料與介面
			self.csv_data = []
			for item in self.csv_tree.get_children():
				self.csv_tree.delete(item)
			self.csv_tree.selection_remove(self.csv_tree.selection())
			self.csv_selection_label.config(text="已選取: 0 個元素")
			self.csv_name_label.config(text="未載入", foreground="gray")
			self.update_match_status()

		except Exception as e:
			messagebox.showerror("錯誤", f"寫入失敗：{str(e)}")

	def save_config(self):
		"""保存配置"""
		new_config_name = self.new_config_var.get().strip()
		current_config_name = self.config_var.get().strip()
		config_name = new_config_name if new_config_name else current_config_name

		if not config_name:
			messagebox.showwarning("警告", "請輸入配置名稱或選擇現有配置")
			return

		# 只保存 element
		selected_items = self.csv_tree.selection()
		selected_elements = []
		for item_id in selected_items:
			item_index = self.csv_tree.index(item_id)
			if item_index < len(self.csv_data):
				csv_row = self.csv_data[item_index]
				selected_elements.append(csv_row.get('Element', ''))

		# 判斷是否為使用選取範圍模式（沒有第二關鍵字但有空格）
		second_keyword = self.field_var.get().strip()
		use_selection_mode = (not second_keyword and len(self.empty_cells) > 0)

		config_data = {
			'first_keyword': self.first_keyword_var.get(),  # 保存第一個關鍵字
			'field_name': self.field_var.get(),  # 第二個關鍵字
			'selected_elements': selected_elements,
			'use_selection_mode': use_selection_mode  # 標記是否為選取範圍模式
		}

		self.field_mappings[config_name] = config_data

		try:
			with open(FIELD_MAPPING_PATH, 'w', encoding='utf-8') as f:
				json.dump(self.field_mappings, f, ensure_ascii=False, indent=2)

			self.update_config_list()
			self.config_var.set(config_name)
			self.new_config_var.set('')

			# 提示用戶保存成功及模式
			mode_text = "選取範圍模式" if use_selection_mode else "獲取空格位置模式"
			messagebox.showinfo("成功", f"配置 '{config_name}' 已保存\n模式: {mode_text}")
		except Exception as e:
			messagebox.showerror("錯誤", f"保存配置失敗：{str(e)}")

	def load_config(self):
		"""套用配置"""
		config_name = self.config_var.get().strip()
		if not config_name or config_name not in self.field_mappings:
			messagebox.showwarning("警告", "請選擇有效的配置")
			return

		try:
			config_data = self.field_mappings[config_name]

			# 加載第一個關鍵字（如果存在）
			if 'first_keyword' in config_data:
				self.first_keyword_var.set(config_data['first_keyword'])
			else:
				self.first_keyword_var.set('')  # 舊配置可能沒有這個欄位

			# 加載第二個關鍵字
			self.field_var.set(config_data['field_name'])

			if not self.active_worksheet and not self.excel_sheet:
				messagebox.showwarning("警告", "請先連接Excel，然後重新套用配置")
				return

			# 檢查是否為選取範圍模式
			use_selection_mode = config_data.get('use_selection_mode', False)

			if use_selection_mode:
				# 選取範圍模式：提示用戶手動選取範圍
				messagebox.showinfo("提示",
					f"配置 '{config_name}' 是使用選取範圍模式\n\n"
					"請在Excel中選取要填入的儲存格範圍，\n"
					"然後點擊「使用選取範圍」按鈕")
			else:
				# 獲取空格位置模式：自動掃描
				self.scan_empty_cells()

			# 自動選取CSV元素
			if 'selected_elements' in config_data and self.csv_data:
				self.auto_select_elements(config_data['selected_elements'])
		except Exception as e:
			messagebox.showerror("錯誤", f"套用配置失敗：{str(e)}")

	def load_configs(self):
		"""套用保存的配置"""
		try:
			if os.path.exists(FIELD_MAPPING_PATH):
				with open(FIELD_MAPPING_PATH, 'r', encoding='utf-8') as f:
					self.field_mappings = json.load(f)
		except Exception as e:
			self.field_mappings = {}

		# 更新配置列表
		self.update_config_list()

	def update_config_list(self):
		"""更新配置下拉列表"""
		config_names = list(self.field_mappings.keys())
		self.config_combo['values'] = config_names

		# 不自動選擇配置，讓使用者手動選擇
		# 如果當前選中的配置不存在於列表中，清空選擇
		current_config = self.config_var.get()
		if current_config and current_config not in config_names:
			self.config_var.set('')

	def auto_apply_current_config(self):
		"""載入CSV時自動套用當前選中的配置"""
		try:
			current_config = self.config_var.get().strip()
			if not current_config or current_config not in self.field_mappings:
				return  # 沒有選中配置或配置不存在，自動忽略

			config_data = self.field_mappings[current_config]

			# 設定第一個關鍵字
			if 'first_keyword' in config_data:
				self.first_keyword_var.set(config_data['first_keyword'])
			else:
				self.first_keyword_var.set('')

			# 設定第二個關鍵字
			if 'field_name' in config_data:
				self.field_var.set(config_data['field_name'])

			# 檢查是否為選取範圍模式
			use_selection_mode = config_data.get('use_selection_mode', False)

			# 如果Excel已連接，且不是選取範圍模式，嘗試獲取空格位置
			if (self.active_worksheet or self.excel_sheet) and not use_selection_mode:
				try:
					self.scan_empty_cells()
				except:
					# 掃描失敗時自動忽略
					pass

			# 自動選中對應的CSV元素
			if 'selected_elements' in config_data and self.csv_data:
				self.auto_select_elements(config_data['selected_elements'])

		except Exception:
			# 任何錯誤都自動忽略，不打擾使用者
			pass

	def auto_select_elements(self, selected_elements):
		"""根據配置自動選中對應的CSV元素"""
		if not self.csv_data or not selected_elements:
			return

		# 清除當前選擇
		self.csv_tree.selection_remove(self.csv_tree.selection())

		selected_items = []

		# 只比對 element
		for config_element_name in selected_elements:
			for i, csv_row in enumerate(self.csv_data):
				csv_element = csv_row.get('Element', '')
				if csv_element == config_element_name:
					children = self.csv_tree.get_children()
					if i < len(children):
						item_id = children[i]
						selected_items.append(item_id)
						break

		if selected_items:
			self.csv_tree.selection_set(selected_items)
			self.update_selection_info()


	def delete_config(self):
		"""刪除配置"""
		config_name = self.config_var.get().strip()
		if not config_name or config_name not in self.field_mappings:
			messagebox.showwarning("警告", "請選擇要刪除的配置")
			return

		# 確認刪除
		result = messagebox.askyesno("確認刪除",
			f"確定要刪除配置 '{config_name}' 嗎？\n此操作無法撤銷。")

		if not result:
			return

		try:
			# 從記憶體中刪除
			del self.field_mappings[config_name]

			# 保存到檔案
			with open(FIELD_MAPPING_PATH, 'w', encoding='utf-8') as f:
				json.dump(self.field_mappings, f, ensure_ascii=False, indent=2)

			# 清空當前選擇
			self.config_var.set('')

			self.update_config_list()
		except Exception as e:
			messagebox.showerror("錯誤", f"刪除配置失敗：{str(e)}")

	def run(self):
		"""運行程序"""
		self.root.mainloop()

if __name__ == "__main__":
	app = SmartExcelMapper()
	app.run()